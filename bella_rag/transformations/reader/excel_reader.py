from io import IOBase, BytesIO
from typing import List

import pandas as pd
from openpyxl.reader.excel import load_workbook

from bella_rag.schema.document import Document, ExcelDocument
from bella_rag.transformations.reader.base import BaseReader


class ExcelReader(BaseReader):

    def load_data(self, stream: IOBase) -> List[Document]:
        """
        reader负责将数据都先读出来，不做任何处理，parse的时候在做处理
        """
        documents = []
        # 1.加载 Excel 文件
        bytes_stream = BytesIO(stream.read())

        # 使用 load_workbook 读取文件流
        workbook = load_workbook(filename=bytes_stream)
        # 遍历每个工作表
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            text = ""

            # 创建一个空列表存储每行数据
            data = []

            # 逐行读取数据
            for row in sheet.iter_rows(values_only=True):
                data.append(list(row))

            # 填充合并单元格
            for merged_cells in sheet.merged_cells.ranges:
                # 获取合并区域左上角单元格的值
                start_cell = merged_cells.start_cell
                merged_value = sheet[start_cell.coordinate].value

                # 填充合并区域的每个单元格
                min_row, min_col, max_row, max_col = merged_cells.min_row - 1, merged_cells.min_col - 1, merged_cells.max_row - 1, merged_cells.max_col - 1
                for row in range(min_row, max_row + 1):
                    for col in range(min_col, max_col + 1):
                        # 填充 data 列表中的相应位置
                        data[row][col] = merged_value

            # 构建text内容：从所有单元格数据中提取
            text_parts = []
            for row_data in data:
                for cell_value in row_data:
                    if cell_value is not None and str(cell_value).strip():
                        text_parts.append(str(cell_value).strip())
            text = " ".join(text_parts)

            if not data:
                # 空sheet则直接跳过
                continue

            # 将填充后的数据转换为 DataFrame
            df = pd.DataFrame(data, columns=[cell for cell in data[0]])  # 用填充后的 data 第一行作为列名

            # 删除全空列
            df.dropna(axis=1, how='all', inplace=True)

            # 表头和无效行列处理
            def col_gen(t_d):
                col_num = t_d.shape[1]
                # 判断表头是否合理
                not_none_num = len([col for col in t_d.columns if col is not None])

                if (not_none_num == 0) or (not_none_num * 1.0 / col_num < 0.5):
                    chg_flag = 0
                    for i in range(t_d.shape[0]):
                        # 这里的修正
                        not_none_num = len([col for col in t_d.values.tolist()[i] if col is not None])
                        if (not_none_num == 0) or (not_none_num * 1.0 / col_num < 0.5):
                            continue
                        chg_flag = 1
                        break

                    if chg_flag == 1:
                        t_d.columns = t_d.values.tolist()[i]  # 这里的修正
                        t_d = t_d.drop(index=range(i + 1)).reset_index(drop=True)

                return t_d

            df_new = col_gen(df)
            # 将 DataFrame 存储到元组中，以工作表名称为键
            documents.append(ExcelDocument(sheet=(sheet_name, df_new), text=text))
        return documents
